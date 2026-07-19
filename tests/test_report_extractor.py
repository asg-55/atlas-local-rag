import io
import unittest
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook
from PIL import Image

from rag_assistant.report_extractor import (
    JOURNAL_COLUMNS,
    REPORT_COLUMNS,
    export_reports_xlsx,
    extract_report_page,
)


class ReportExtractorTests(unittest.TestCase):
    @patch("rag_assistant.report_extractor._detect_line", return_value=("Б", 0.95))
    @patch("rag_assistant.report_extractor._read_value_cell")
    @patch("rag_assistant.report_extractor._classify_template_tables")
    @patch("rag_assistant.report_extractor._detect_table_regions", return_value=[])
    @patch("rag_assistant.report_extractor._prepare_ocr_image", return_value=None)
    def test_partial_page_keeps_available_data_and_warnings(self, _prepare, _detect, classify, read_cell, _line):
        rows = [[(0, index, 10, index + 1), (10, index, 20, index + 1)] for index in range(7)]
        classify.return_value = {
            "general": rows,
            "catalyst": [],
            "teal": [],
            "donor": [],
            "heptane": [],
            "c3": [],
            "journal": [],
        }

        def read_value(_image, _box, kind="number", value_range=None):
            return {"date": "19.07.2025", "batch": "BCNX-24-CO"}.get(kind, "1"), 0.90

        read_cell.side_effect = read_value
        report, journal, warnings = extract_report_page(Image.new("RGB", (100, 100)), "partial.pdf", 1)

        self.assertEqual(1, report["№ барабана"])
        self.assertEqual("Проверить", report["Статус"])
        self.assertEqual([], journal)
        self.assertTrue(any("Не найдена таблица: c3" in warning for warning in warnings))
        self.assertTrue(any("Не найдены заполненные строки журнала" in warning for warning in warnings))

    @patch("rag_assistant.report_extractor._detect_line", return_value=("Б", 0.95))
    @patch("rag_assistant.report_extractor._read_value_cell")
    @patch("rag_assistant.report_extractor._classify_template_tables")
    @patch("rag_assistant.report_extractor._detect_table_regions", return_value=[])
    @patch("rag_assistant.report_extractor._prepare_ocr_image", return_value=None)
    def test_extracts_all_five_journal_columns(self, _prepare, _detect, classify, read_cell, _line):
        next_box = 0
        values = {}

        def box(value):
            nonlocal next_box
            next_box += 1
            token = (next_box, 0, next_box + 1, 1)
            values[token] = str(value)
            return token

        scalar_values = {
            "general": ["19.07.2025", 34, "BCNX-24-CO", 2, 4, 55, 60.00],
            "catalyst": [40, 16.7],
            "teal": [6, 100.2, 11424, 166],
            "donor": [0.30, 5.01, 941.96, 12.56],
            "heptane": [8000, 10000, 108, 150, 9400],
            "c3": [4, 160, 6, 80, 20],
        }
        tables = {name: [[box("label"), box(value)] for value in group] for name, group in scalar_values.items()}
        tables["journal"] = [
            [box("18:10"), box(20), box(0.02), box(19.65), box(6.54)],
            [box("0:00"), box(0), box(0), box(0), box(0)],
        ]
        classify.return_value = tables

        def read_value(_image, token, kind="number", value_range=None):
            value = values[token]
            if value == "label":
                return None, 0.0
            return value, 0.90

        read_cell.side_effect = read_value
        report, journal, warnings = extract_report_page(Image.new("RGB", (100, 100)), "batch.pdf", 1)

        self.assertEqual("Готово", report["Статус"])
        self.assertEqual("Б", report["Линия"])
        self.assertEqual(19.65, journal[0]["P в D2B, кПа"])
        self.assertEqual(6.54, journal[0]["Температура в DZB, °C"])
        self.assertEqual(1, len(journal))
        self.assertEqual([], warnings)

    def test_exports_three_analysis_sheets(self):
        report = {column: None for column in REPORT_COLUMNS}
        report.update({"Файл": "batch.pdf", "Страница": 1, "Статус": "Готово", "Уверенность OCR": 0.82, "Полнота": 1.0})
        journal = {column: None for column in JOURNAL_COLUMNS}
        journal.update({"Файл": "batch.pdf", "Страница": 1, "Время": "18:10", "P в D2B, кПа": 19.65})
        content = export_reports_xlsx(
            pd.DataFrame([report], columns=REPORT_COLUMNS),
            pd.DataFrame([journal], columns=JOURNAL_COLUMNS),
            pd.DataFrame([], columns=["Файл", "Страница", "Предупреждение"]),
        )
        workbook = load_workbook(io.BytesIO(content), data_only=False)

        self.assertEqual(["Отчеты", "Журнал", "Контроль"], workbook.sheetnames)
        self.assertEqual("batch.pdf", workbook["Отчеты"]["A2"].value)
        self.assertEqual(19.65, workbook["Журнал"]["F2"].value)


if __name__ == "__main__":
    unittest.main()
