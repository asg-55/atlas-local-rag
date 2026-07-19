from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Callable

import fitz
import cv2
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image

from .parsers import _detect_table_regions, _ocr_reader, _prepare_ocr_image


REPORT_COLUMNS = [
    "Файл", "Страница", "Линия", "Статус", "Уверенность OCR", "Полнота",
    "Дата", "№ барабана", "№ партии катализатора", "Ti в партии, вес. %",
    "Концентрация катализатора в D2B, г/л", "Концентрация TEAL в D5Z, г/л Hp",
    "Концентрация донора C в D7B, г/л Hp", "Катализатор, кг", "Ti, моль",
    "Al/Ti, моль/моль", "Всего Al, моль", "Чистый TEAL, г", "Раствор TEAL, кг",
    "Si/Ti, моль/моль", "Всего Si, моль", "Чистый донор C, г", "Раствор донора C, кг",
    "Общий объем D2B 80%, л", "Hp по плановой концентрации, л",
    "Hp с TEAL и донором C, л", "Hp для промывки барабана, л", "Исходный поток Hp, л",
    "Соотношение C3:катализатор, кг/кг", "Всего C3, кг", "Плановое время, ч",
    "Базовая скорость C3, кг/ч", "Начальная скорость C3, кг/ч",
]

JOURNAL_COLUMNS = [
    "Файл", "Страница", "Время", "Подача C3, кг/ч", "Общая подача C3, кг",
    "P в D2B, кПа", "Температура в DZB, °C",
]

SCALAR_TABLE_FIELDS = {
    2: [
        "Дата", "№ барабана", "№ партии катализатора", "Ti в партии, вес. %",
        "Концентрация катализатора в D2B, г/л", "Концентрация TEAL в D5Z, г/л Hp",
        "Концентрация донора C в D7B, г/л Hp",
    ],
    3: ["Катализатор, кг", "Ti, моль"],
    4: ["Al/Ti, моль/моль", "Всего Al, моль", "Чистый TEAL, г", "Раствор TEAL, кг"],
    5: ["Si/Ti, моль/моль", "Всего Si, моль", "Чистый донор C, г", "Раствор донора C, кг"],
    6: [
        "Общий объем D2B 80%, л", "Hp по плановой концентрации, л",
        "Hp с TEAL и донором C, л", "Hp для промывки барабана, л", "Исходный поток Hp, л",
    ],
    7: [
        "Соотношение C3:катализатор, кг/кг", "Всего C3, кг", "Плановое время, ч",
        "Базовая скорость C3, кг/ч", "Начальная скорость C3, кг/ч",
    ],
}


def pdf_page_count(content: bytes) -> int:
    with fitz.open(stream=content, filetype="pdf") as pdf:
        return len(pdf)


def render_pdf_page(content: bytes, page_number: int, dpi: int = 130) -> bytes:
    with fitz.open(stream=content, filetype="pdf") as pdf:
        page = pdf[page_number - 1]
        return page.get_pixmap(dpi=dpi, alpha=False).tobytes("png")


def _cells(text: str) -> list[list[str]]:
    return [[cell.strip() for cell in row.split("|")] for row in text.splitlines() if row.strip()]


def _last_value(row: list[str]) -> str | None:
    values = [value.strip() for value in row if value.strip()]
    return values[-1] if len(values) >= 2 else None


def _typed_value(value: str | None, field: str):
    if value is None:
        return None
    value = value.strip()
    if not value:
        return None
    if field == "Дата":
        for fmt in ("%d.%m.%Y", "%d.%m.%y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
        return value
    if field in {"№ партии катализатора"}:
        return value
    normalized = value.replace(" ", "").replace(",", ".")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", normalized)
    if not match:
        return value
    number = float(match.group())
    return int(number) if number.is_integer() else number


def _data_rows(table_text: str) -> list[list[str]]:
    rows = _cells(table_text)
    result = []
    for row in rows:
        value = _last_value(row)
        if value is None:
            continue
        if not re.search(r"\d", value):
            continue
        result.append(row)
    return result


def _table_bounds(table) -> tuple[int, int, int, int]:
    boxes = [box for row in table for box in row]
    return (
        min(box[0] for box in boxes), min(box[1] for box in boxes),
        max(box[2] for box in boxes), max(box[3] for box in boxes),
    )


def _classify_template_tables(tables) -> dict[str, list]:
    """Classify tables by grid shape, not page pixels or contour order."""
    journal_candidates = [
        table for table in tables
        if len(table) >= 10 and max(len(row) for row in table) >= 4
    ]
    journal = max(journal_candidates, key=len) if journal_candidates else []
    scalar = [table for table in tables if table is not journal and max(len(row) for row in table) <= 2]
    scalar.sort(key=lambda table: (_table_bounds(table)[1], _table_bounds(table)[0]))
    general = max(scalar, key=len) if scalar else []
    remaining = [table for table in scalar if table is not general]
    remaining.sort(key=lambda table: _table_bounds(table)[1])
    result = {
        "journal": journal,
        "general": general,
    }
    for index, name in enumerate(("catalyst", "teal", "donor", "heptane", "c3")):
        result[name] = remaining[index] if index < len(remaining) else []
    return result


def _cell_variants(image_array: np.ndarray, box) -> list[np.ndarray]:
    x1, y1, x2, y2 = box
    pad_x = max(2, int((x2 - x1) * 0.025))
    pad_y = max(2, int((y2 - y1) * 0.08))
    crop = image_array[max(0, y1 + pad_y):max(0, y2 - pad_y), max(0, x1 + pad_x):max(0, x2 - pad_x)]
    if crop.size == 0:
        return []
    scale = max(2.0, min(4.0, 90 / max(1, crop.shape[0])))
    enlarged = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(enlarged, cv2.COLOR_RGB2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8)).apply(gray)
    _, otsu = cv2.threshold(clahe, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    adaptive = cv2.adaptiveThreshold(clahe, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 9)
    return [enlarged, cv2.cvtColor(otsu, cv2.COLOR_GRAY2RGB), cv2.cvtColor(adaptive, cv2.COLOR_GRAY2RGB)]


def _normalize_candidate(text: str, kind: str) -> str:
    value = re.sub(r"\s+", "", text).strip("|_;")
    if kind == "time":
        value = value.replace(".", ":").replace(",", ":")
    elif kind == "date":
        value = value.replace(":", ".").replace(",", ".")
    elif kind == "number":
        value = value.replace(",", ".")
    elif kind == "batch":
        value = value.upper().replace("—", "-").replace("_", "-")
    return value


def _valid_candidate(value: str, kind: str, value_range: tuple[float, float] | None = None) -> bool:
    if kind == "time":
        if not re.fullmatch(r"\d{1,2}:\d{2}", value):
            return False
        hour, minute = map(int, value.split(":"))
        return 0 <= hour <= 23 and 0 <= minute <= 59
    if kind == "date":
        return bool(re.fullmatch(r"\d{1,2}\.\d{1,2}\.\d{2,4}", value))
    if kind == "batch":
        return bool(re.fullmatch(r"[A-Z0-9-]{4,30}", value) and any(ch.isdigit() for ch in value))
    if kind == "number":
        if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", value):
            return False
        if value_range:
            number = float(value)
            return value_range[0] <= number <= value_range[1]
        return True
    return bool(value)


def _read_value_cell(
    image_array: np.ndarray,
    box,
    kind: str = "number",
    value_range: tuple[float, float] | None = None,
) -> tuple[str | None, float]:
    allowlist = {
        "number": "0123456789.,-+",
        "time": "0123456789.:,",
        "date": "0123456789.:,",
        "batch": "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_",
    }.get(kind)
    candidates: list[tuple[str, float]] = []
    for variant_index, variant in enumerate(_cell_variants(image_array, box)):
        raw = _ocr_reader().readtext(
            variant,
            detail=1,
            paragraph=False,
            allowlist=allowlist,
            text_threshold=0.45 if variant_index == 0 else 0.3,
            low_text=0.2,
            link_threshold=0.25,
        )
        if not raw:
            continue
        text = "".join(str(item[1]) for item in raw)
        confidence = float(np.mean([float(item[2]) for item in raw]))
        value = _normalize_candidate(text, kind)
        if _valid_candidate(value, kind, value_range):
            candidates.append((value, confidence))
        if candidates and candidates[-1][1] >= 0.93:
            break
    if not candidates:
        return None, 0.0
    # Agreement between preprocessing variants is more reliable than one high score.
    frequencies = {value: sum(1 for candidate, _ in candidates if candidate == value) for value, _ in candidates}
    return max(candidates, key=lambda item: (frequencies[item[0]], item[1]))


def _value_rows(table, expected_count: int) -> list:
    rows = [row for row in table if len(row) >= 2]
    return rows[-expected_count:]


def _value_boxes(image_array: np.ndarray, table, expected_count: int) -> list[tuple[int, int, int, int]]:
    """Rebuild value cells from the table grid when an individual contour is missed."""
    rows = [row for row in table if len(row) >= 2]
    fallback = [row[-1] for row in rows[-expected_count:]]
    if not rows or image_array is None:
        return fallback

    x1, y1, x2, y2 = _table_bounds(table)
    gray = cv2.cvtColor(image_array[y1:y2 + 1, x1:x2 + 1], cv2.COLOR_RGB2GRAY)
    binary = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)[1]
    # Preserve long grid strokes while removing letters and digits from the projection.
    horizontal = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (max(30, binary.shape[1] // 5), 1)),
    )
    density = np.mean(horizontal > 0, axis=1)
    line_indices = np.where(density >= 0.12)[0]
    if not len(line_indices):
        return fallback

    clusters: list[list[int]] = [[int(line_indices[0])]]
    for index in line_indices[1:]:
        if int(index) <= clusters[-1][-1] + 1:
            clusters[-1].append(int(index))
        else:
            clusters.append([int(index)])
    boundaries = [y1 + int(round(float(np.mean(cluster)))) for cluster in clusters]
    intervals = [
        (top, bottom)
        for top, bottom in zip(boundaries, boundaries[1:])
        if 18 <= bottom - top <= 160
    ]
    if len(intervals) < expected_count:
        return fallback

    value_x1 = int(round(float(np.median([row[-1][0] for row in rows]))))
    value_x2 = max(row[-1][2] for row in rows)
    return [(value_x1, top, value_x2, bottom) for top, bottom in intervals[-expected_count:]]


def _detect_line(image_array: np.ndarray) -> tuple[str | None, float]:
    header = image_array[: max(180, int(image_array.shape[0] * 0.14)), :]
    raw = _ocr_reader().readtext(header, detail=1, paragraph=False)
    confidence = float(np.mean([float(item[2]) for item in raw])) if raw else 0.0
    for item in raw:
        text = str(item[1]).upper()
        match = re.search(r"ЛИНИ[ИЯ]\s*([AАBБ])", text)
        if match:
            return "А" if match.group(1) in {"A", "А"} else "Б", float(item[2])
    return None, confidence


def _parse_journal(table_text: str, filename: str, page_number: int) -> list[dict]:
    journal = []
    for row in _cells(table_text):
        values = [value for value in row if value]
        if len(values) < 5 or not re.fullmatch(r"\d{1,2}[.:]\d{2}", values[0]):
            continue
        numeric = []
        for value in values[1:5]:
            match = re.search(r"[-+]?\d+(?:[.,]\d+)?", value.replace(" ", ""))
            numeric.append(float(match.group().replace(",", ".")) if match else None)
        if all(value in (None, 0.0) for value in numeric):
            continue
        journal.append(dict(zip(JOURNAL_COLUMNS, [filename, page_number, values[0].replace(".", ":"), *numeric])))
    return journal


def extract_report_page(image: Image.Image, filename: str, page_number: int) -> tuple[dict, list[dict], list[str]]:
    image_array = _prepare_ocr_image(image)
    tables = _classify_template_tables(_detect_table_regions(image_array))
    report = {column: None for column in REPORT_COLUMNS}
    report.update({"Файл": filename, "Страница": page_number})
    warnings: list[str] = []
    confidences: list[float] = []

    line, line_confidence = _detect_line(image_array)
    report["Линия"] = line
    if line:
        confidences.append(line_confidence)
    else:
        warnings.append("Не удалось определить линию А/Б в шапке")

    field_groups = {
        "general": SCALAR_TABLE_FIELDS[2],
        "catalyst": SCALAR_TABLE_FIELDS[3],
        "teal": SCALAR_TABLE_FIELDS[4],
        "donor": SCALAR_TABLE_FIELDS[5],
        "heptane": SCALAR_TABLE_FIELDS[6],
        "c3": SCALAR_TABLE_FIELDS[7],
    }
    for table_name, fields in field_groups.items():
        if not tables[table_name]:
            warnings.append(f"Не найдена таблица: {table_name}; поля оставлены пустыми")
            continue
        value_boxes = _value_boxes(image_array, tables[table_name], len(fields))
        if len(value_boxes) < len(fields):
            warnings.append(f"{table_name}: найдено {len(value_boxes)} из {len(fields)} строк значений")
        for field, value_box in zip(fields, value_boxes):
            kind = "date" if field == "Дата" else "batch" if field == "№ партии катализатора" else "number"
            value_range = (0, 10000) if field == "№ барабана" else (0, 10_000_000) if kind == "number" else None
            value, confidence = _read_value_cell(image_array, value_box, kind, value_range)
            report[field] = _typed_value(value, field)
            if value is not None:
                confidences.append(confidence)
                if confidence < 0.6:
                    warnings.append(f"Низкая уверенность: {field} = {value} ({confidence:.0%})")
            else:
                warnings.append(f"Не распознано поле: {field}")

    journal: list[dict] = []
    for row in tables.get("journal", []):
        if len(row) < 5:
            continue
        time_value, time_confidence = _read_value_cell(image_array, row[0], "time")
        if time_value is None:
            continue
        ranges = [(0, 500), (0, 1000), (0, 1000), (-100, 500)]
        numeric_values: list[float | None] = []
        row_confidences = [time_confidence]
        for box, value_range in zip(row[1:5], ranges):
            raw_value, confidence = _read_value_cell(image_array, box, "number", value_range)
            numeric_values.append(float(raw_value) if raw_value is not None else None)
            if raw_value is not None:
                row_confidences.append(confidence)
        if all(value in (None, 0.0) for value in numeric_values):
            continue
        if any(value is None for value in numeric_values):
            warnings.append(f"Журнал {time_value}: распознаны не все четыре значения")
        confidences.extend(row_confidences)
        journal.append(dict(zip(JOURNAL_COLUMNS, [filename, page_number, time_value, *numeric_values])))
    if not journal:
        warnings.append("Не найдены заполненные строки журнала")
    confidence = float(np.mean(confidences)) if confidences else 0.0
    data_fields = REPORT_COLUMNS[6:]
    filled = sum(report.get(field) not in (None, "") for field in data_fields)
    completeness = filled / max(1, len(data_fields))
    report["Уверенность OCR"] = round(confidence, 3)
    report["Полнота"] = round(completeness, 3)
    report["Статус"] = "Готово" if completeness >= 0.95 and journal and confidence >= 0.6 and line else "Проверить"
    if confidence < 0.6:
        warnings.append(f"Низкая средняя уверенность OCR: {confidence:.0%}")
    return report, journal, warnings


def extract_batch_pdf(
    content: bytes,
    filename: str,
    start_page: int = 1,
    end_page: int | None = None,
    dpi: int = 220,
    progress: Callable[[int, int, str], None] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    reports: list[dict] = []
    journal: list[dict] = []
    quality: list[dict] = []
    with fitz.open(stream=content, filetype="pdf") as pdf:
        last_page = min(end_page or len(pdf), len(pdf))
        pages = list(range(max(1, start_page), last_page + 1))
        for index, page_number in enumerate(pages, start=1):
            if progress:
                progress(index - 1, len(pages), f"Страница {page_number} из {last_page}")
            page = pdf[page_number - 1]
            pix = page.get_pixmap(dpi=dpi, alpha=False)
            image = Image.open(io.BytesIO(pix.tobytes("png")))
            try:
                report, rows, warnings = extract_report_page(image, filename, page_number)
                reports.append(report)
                journal.extend(rows)
                quality.extend(
                    {"Файл": filename, "Страница": page_number, "Предупреждение": warning}
                    for warning in warnings
                )
            except Exception as exc:
                reports.append({**{column: None for column in REPORT_COLUMNS}, "Файл": filename, "Страница": page_number, "Статус": "Ошибка"})
                quality.append({"Файл": filename, "Страница": page_number, "Предупреждение": str(exc)})
        if progress:
            progress(len(pages), len(pages), "Обработка завершена")
    return (
        pd.DataFrame(reports, columns=REPORT_COLUMNS),
        pd.DataFrame(journal, columns=JOURNAL_COLUMNS),
        pd.DataFrame(quality, columns=["Файл", "Страница", "Предупреждение"]),
    )


def _style_sheet(sheet, table_name: str) -> None:
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = sheet.dimensions
    if sheet.max_row >= 2 and sheet.max_column >= 1:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        sheet.add_table(table)
    header_fill = PatternFill("solid", fgColor="242424")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    sheet.row_dimensions[1].height = 38
    for column in sheet.columns:
        letter = column[0].column_letter
        width = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[letter].width = min(max(width + 2, 11), 34)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def export_reports_xlsx(reports: pd.DataFrame, journal: pd.DataFrame, quality: pd.DataFrame) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, dataframe, table_name in (
        ("Отчеты", reports, "ReportsTable"),
        ("Журнал", journal, "JournalTable"),
        ("Контроль", quality, "QualityTable"),
    ):
        sheet = workbook.create_sheet(title)
        sheet.append(list(dataframe.columns))
        for row in dataframe.itertuples(index=False, name=None):
            sheet.append([None if pd.isna(value) else value for value in row])
        _style_sheet(sheet, table_name)
    reports_sheet = workbook["Отчеты"]
    header_index = {cell.value: cell.column for cell in reports_sheet[1]}
    for name in ("Уверенность OCR", "Полнота"):
        column = header_index.get(name)
        if column:
            for cell in reports_sheet.iter_cols(min_col=column, max_col=column, min_row=2, max_row=reports_sheet.max_row):
                for item in cell:
                    item.number_format = "0.0%"
            letter = reports_sheet.cell(1, column).column_letter
            reports_sheet.conditional_formatting.add(
                f"{letter}2:{letter}{reports_sheet.max_row}",
                CellIsRule(operator="lessThan", formula=["0.7"], fill=PatternFill("solid", fgColor="FECACA")),
            )
    date_column = header_index.get("Дата")
    if date_column:
        for row in range(2, reports_sheet.max_row + 1):
            reports_sheet.cell(row, date_column).number_format = "dd.mm.yyyy"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
